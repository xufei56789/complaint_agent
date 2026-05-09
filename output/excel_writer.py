"""
output/excel_writer.py
Excel 结果文件生成

write_excel() 写入三个标签页：
  - 四要素：工单编号 / 时间 / 地点 / 人物 / 事件
  - 四级分类：工单编号 / 一级 / 二级 / 三级 / 四级分类
  - 群诉：群诉编号 / 工单编号 / 涉及点位 / 共性问题 / 涉及工单数 / 事件摘要
"""
from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_FILE

# ── 配色 ─────────────────────────────────────────────────────
_C = {
    "hdr_blue":   "1F4E79",
    "hdr_green":  "375623",
    "hdr_orange": "7F3F00",
    "mid_blue":   "2E75B6",
    "mid_green":  "4EAC3D",
    "mid_orange": "C55A11",
    "row_odd":    "F2F7FB",
    "highlight":  "FFF2CC",
}

_L1_COLORS = {
    "城市管理": "D9E8F5",
    "住房建设": "D5E8D4",
    "市场监管": "FFF2CC",
    "教育卫生": "F8CECC",
    "社会民生": "E1D5E7",
    "其他诉求": "F5F5F5",
}

_THIN = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)


# ── 样式辅助 ──────────────────────────────────────────────────
def _hdr(cell, text: str, bg: str) -> None:
    cell.value = text
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _THIN


def _title(cell, text: str, bg: str, size: int = 14) -> None:
    cell.value = text
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _cell(cell, value, bg: str | None = None, wrap: bool = False, bold: bool = False) -> None:
    cell.value     = value
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap, horizontal="left")
    cell.border    = _THIN
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


# ============================================================
# Sheet 1：四要素
# ============================================================
def _write_four_elements(wb: Workbook, data: List[Dict]) -> None:
    ws = wb.active
    ws.title = "四要素"

    headers   = ["工单编号", "时间", "地点", "人物", "事件"]
    col_widths = [12, 22, 30, 28, 60]

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A1:E1")
    _title(ws["A1"], "工单四要素分析表", _C["hdr_blue"])

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws.cell(2, ci), h, _C["mid_blue"])
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(data, 3):
        ws.row_dimensions[ri].height = 45
        bg = _C["row_odd"] if ri % 2 == 1 else None
        for ci, key in enumerate(headers, 1):
            _cell(ws.cell(ri, ci), row.get(key, ""), bg=bg, wrap=(ci == 5))

    ws.freeze_panes = "A3"


# ============================================================
# Sheet 2：四级分类
# ============================================================
def _write_classifications(wb: Workbook, data: List[Dict]) -> None:
    ws = wb.create_sheet("四级分类")

    headers    = ["工单编号", "一级分类", "二级分类", "三级分类", "四级分类"]
    col_widths = [12, 15, 18, 20, 22]

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A1:E1")
    _title(ws["A1"], "工单四级分类表", _C["hdr_green"])

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws.cell(2, ci), h, _C["mid_green"])
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(data, 3):
        ws.row_dimensions[ri].height = 22
        bg = _L1_COLORS.get(row.get("一级分类", ""), "FFFFFF")
        for ci, key in enumerate(headers, 1):
            c = ws.cell(ri, ci)
            c.value     = row.get(key, "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(vertical="center", horizontal="center" if ci <= 4 else "left")
            c.border    = _THIN

    ws.freeze_panes = "A3"


# ============================================================
# Sheet 3：群诉
# ============================================================
def _write_group_complaints(wb: Workbook, data: List[Dict]) -> None:
    ws = wb.create_sheet("群诉")

    if not data:
        ws["A1"] = "暂未检测到明确群诉工单"
        return

    headers    = ["群诉编号", "工单编号", "涉及点位", "共性问题", "涉及工单数", "事件摘要"]
    col_widths = [14, 12, 28, 18, 12, 55]

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A1:F1")
    _title(ws["A1"], "群诉工单分析表（同点位、同问题的多工单聚合）", _C["hdr_orange"], size=12)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws.cell(2, ci), h, _C["mid_orange"])
        ws.column_dimensions[get_column_letter(ci)].width = w

    prev_group = None
    for ri, row in enumerate(data, 3):
        ws.row_dimensions[ri].height = 40
        cur_group = row.get("群诉编号", "")
        is_new    = cur_group != prev_group
        prev_group = cur_group
        bg = _C["highlight"] if is_new else None
        for ci, key in enumerate(headers, 1):
            _cell(ws.cell(ri, ci), row.get(key, ""), bg=bg, wrap=(ci == 6), bold=is_new)

    ws.freeze_panes = "A3"


# ============================================================
# 主入口
# ============================================================
def write_excel(
    four_elements:    List[Dict],
    classifications:  List[Dict],
    group_complaints: List[Dict],
    output_path: str | Path | None = None,
) -> str:
    """
    生成结果 Excel，返回保存路径字符串。
    output_path 默认取 config/settings.py 中的 OUTPUT_FILE。
    """
    path = Path(output_path) if output_path else OUTPUT_FILE
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_four_elements(wb, four_elements)
    _write_classifications(wb, classifications)
    _write_group_complaints(wb, group_complaints)

    wb.save(str(path))
    return str(path)
